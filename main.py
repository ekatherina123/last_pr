from bs4 import BeautifulSoup
from fake_user_agent import user_agent
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

import time
from datetime import datetime, timedelta
from get_link_on_case import scrap_inf
from global_names import (
    PLAINTIFFS,
    DEFENDANTS,
    THIRDS,
    OTHERS,
    INN,
    DATE,
    NUMBERS_CASE,
    ESSENCE_OF_CASE,
    COURTS,
)
from print_in_exel import print_in_excel_func

fn = "результаты_парсинга.xlsx"
wb = load_workbook(fn)
ws = wb["Лист1"]

# очистка листа
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.value = None

headers = [
    "Номер дела",
    "Дата",
    "Суд",
    "Истец/ИНН",
    "Ответчик/ИНН",
    "Третьи лица",
    "Иные лица",
    "Суть дела",
]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True, size=12)  # Делает текст жирным
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

wb.save(fn)
wb.close()

useragent = user_agent()
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--disable-blink-features=AutomationControlled")


driver = webdriver.Chrome(options=options)
URL = "https://kad.arbitr.ru"

driver.maximize_window()
driver.get(URL)
time.sleep(5)

# Начальная дата
start_date = datetime.strptime("01.06.2024", "%d.%m.%Y")
end_date = datetime.strptime("30.06.2024", "%d.%m.%Y")

current_date = start_date

while current_date < end_date:
    data_first = current_date.strftime("%d.%m.%Y")
    data_second = (current_date + timedelta(days=1)).strftime("%d.%m.%Y")
    # Ввод данных на страницу
    data_input = driver.find_elements(
        By.CSS_SELECTOR, 'input[class="anyway_position_top g-ph"]'
    )
    data_first_input = data_input[0]
    data_second_input = data_input[1]

    data_first_input.click()
    time.sleep(3)
    data_first_input.clear()
    time.sleep(1)
    data_first_input.click()
    time.sleep(1)
    data_first_input.send_keys(data_first)
    time.sleep(1)

    data_second_input.click()
    time.sleep(1)
    data_second_input.clear()
    time.sleep(1)
    data_second_input.click()
    time.sleep(1)
    data_second_input.send_keys(data_second)
    time.sleep(3)
    data_second_input.send_keys(Keys.RETURN)
    time.sleep(5)

    while True:
        try:
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, "lxml")

            scrap_inf(soup)

            print_in_excell(
                DATE,
                NUMBERS_CASE,
                INN,
                PLAINTIFFS,
                DEFENDANTS,
                THIRDS,
                OTHERS,
                ESSENCE_OF_CASE,
                COURTS,
                r"C:\Users\KENG69\PycharmProjects\sber_prakt_norm\результаты_парсинга.xlsx",
            )

            time.sleep(6)

            next_button = driver.find_element(By.CSS_SELECTOR, 'li[class="rarr"]')
            next_button.click()

            time.sleep(6)
        except:
            print("Следующая страница недоступна, переходим к следующей.")
            break

    current_date += timedelta(days=2)

driver.quit()
