from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Side,  Border
from openpyxl.utils import get_column_letter
def print_in_excel_func(
    dates,
    numbers_case,
    inns,
    plaintiffs,
    defendants,
    thirds,
    others,
    essence_of_case,
    courts,
    filename,
):
    work_boook = load_workbook(filename)
    work_sheet = work_boook["Лист1"]

    # Находим следующую строку для записи данных
    next_row = work_sheet.max_row + 1



    for row_index in range(len(numbers_case)):
        # Номер дела
        work_sheet.cell(
            row=next_row + row_index, column=1, value=numbers_case[row_index]
        ).font =Font(size=12)
        work_sheet.cell(row=next_row + row_index, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        work_sheet.cell(row=next_row + row_index, column=1).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Дата
        work_sheet.cell(row=next_row + row_index, column=2, value=dates[row_index]).font = Font(size=12)
        work_sheet.cell(row=next_row + row_index, column=2).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        work_sheet.cell(row=next_row + row_index, column=2).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        # Истец
        work_sheet.cell(
            row=next_row + row_index,
            column=3,
            value=f"{plaintiffs[row_index]} / {inns[row_index][0]}",
        ).font = Font(size=12)
        work_sheet.cell(row=next_row + row_index, column=3).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        work_sheet.cell(row=next_row + row_index, column=3).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Ответчик
        work_sheet.cell(
            row=next_row + row_index,
            column=4,
            value=f"{defendants[row_index]} / {inns[row_index][1]}",
        ).font = Font(size=12)
        work_sheet.cell(row=next_row + row_index, column=4).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        work_sheet.cell(row=next_row + row_index, column=4).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Суд
        work_sheet.cell(
            row=next_row + row_index, column=5, value=courts[row_index]
        ).font = Font(size=12)
        work_sheet.cell(row=next_row + row_index, column=5).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        work_sheet.cell(row=next_row + row_index, column=5).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )


        # Третьи лица
        if thirds[row_index]:
            work_sheet.cell(
                row=next_row + row_index, column=6, value=", ".join(thirds[row_index])
            ).font = Font(size=12)
            work_sheet.cell(row=next_row + row_index, column=6).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            work_sheet.cell(row=next_row + row_index, column=6).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        else:
            work_sheet.cell(row=next_row + row_index, column=6, value="-").font = Font(size=12)
            work_sheet.cell(row=next_row + row_index, column=6).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            work_sheet.cell(row=next_row + row_index, column=6).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Иные лица
        if others[row_index]:
            work_sheet.cell(
                row=next_row + row_index, column=7, value=", ".join(others[row_index])
            ).font = Font(size=12)
            work_sheet.cell(row=next_row + row_index, column=7).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            work_sheet.cell(row=next_row + row_index, column=7).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        else:
            work_sheet.cell(row=next_row + row_index, column=7, value="-").font = Font(size=12)
            work_sheet.cell(row=next_row + row_index, column=7).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            work_sheet.cell(row=next_row + row_index, column=7).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Суть дела
        if essence_of_case[row_index]:
            work_sheet.cell(
                row=next_row + row_index,
                column=8,
                value=", ".join(essence_of_case[row_index]),
            ).font = Font(size=12)
            work_sheet.cell(row=next_row + row_index, column=8).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            work_sheet.cell(row=next_row + row_index, column=8).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        else:
            work_sheet.cell(row=next_row + row_index, column=8, value="-").font = Font(size=12)
            work_sheet.cell(row=next_row + row_index, column=8).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            work_sheet.cell(row=next_row + row_index, column=8).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )


    dates.clear()
    numbers_case.clear()
    inns.clear()
    plaintiffs.clear()
    defendants.clear()
    thirds.clear()
    others.clear()
    essence_of_case.clear()
    courts.clear()

    work_boook.save(filename)
    work_boook.close()
